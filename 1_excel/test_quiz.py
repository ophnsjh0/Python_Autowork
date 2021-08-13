# QUIZ) 여러분은 나보대학의 컴퓨터과 교수님입니다.
# 여러분이 가르치는 과목의 점수 비중은 다음과 같습니다.

# -출석 10
# -퀴즈1 10
# -퀴즈2 10
# -중간고사 20
# -기말고사 30
# -프로젝트 20
# ===============
# -총합 100

# 마지막 수업을 모두 마치고 이번 학기 학생들의 최종 성적을 검토하는 과정에서
# 퀴즈2 문제에 오류를 발견하여 모두 만점 처리를 하기로 하였습니다.
# 현재까지 작성된 최종 성적 데이터를 기준으로 아래와 같이 수정하시오

# 1. 퀴즈2 점수를 10으로 수정
# 2. H열에 총점(SUM 이용), I열에 성적 정보 추가
# - 총점 90 이상 A, 80 이상 B, 70 이상 C, 나머지 D
# 3. 출석이 5 미만인 학생은 총점 상관없이 F

# ※ 최종 파일명 : scores.xlsx

# [현재까지 작성된 최종 성적 데이터]
# 학번, 출석, 귀즈1, 퀴즈2, 중간고사, 기말고사, 프로젝트
# 1,10,8,5,14,26,12
# 2,7,3,7,15,24,18
# 3,9,5,8,8,12,4
# 4,7,8,7,17,21,18
# 5,7,8,7,16,25,15
# 6.3.5.8,8,17,0
# 7,4,9,10,16,27,18
# 8,6,6,6,15,19,17
# 9,10,10,9,19,30,19
# 10,9,8,8,20,25,20

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string
wb = load_workbook("scores.xlsx")
ws = wb.active


def modify():
    for col in ws.iter_rows(min_row=2, max_row=11, min_col=4, max_col=4):
        col[0].value = 10


def attendance():
    for cell in ws.iter_rows(min_row=2, max_row=11, min_col=1, max_col=9):
        # print(cell[1].value)
        if cell[1].value < 5:
            cell[8].value = "F"


def sum():
    row_range = ws[2:11]
    for rows in row_range:
        for cell in rows:
            print(cell.value)


# row_range = ws[2:ws.max_row]  # 2번째 줄부터 마지막 줄까지
# for rows in row_range:
#     for cell in rows:
#         # print(cell.value, end=" ")
#         # print(cell.coordinate, end=" ")
#         xy = coordinate_from_string(cell.coordinate)
#         # print(xy, end=" ")
#         print(xy[0], end="")  # A
#         print(xy[1], end=" ")  # 숫자
#     print()
# ws["A6"] = "=SUM(A4:A5)"  # 30
# modify()
# # attendance()
# sum()


# wb.save("scores_result.xlsx")
