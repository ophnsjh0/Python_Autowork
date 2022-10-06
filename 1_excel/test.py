from openpyxl import load_workbook
wb = load_workbook("test.xlsx")
ws = wb.active


for row in ws.iter_rows(min_row=1):
    for cell in row:
        if cell.value == "add lb vserver AAA":
            cell.value = "test"

wb.save("test2.xlsx")
