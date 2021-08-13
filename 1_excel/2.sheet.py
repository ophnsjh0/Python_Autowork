from openpyxl import Workbook
wb = Workbook()
# wb.active
ws = wb.create_sheet()  # 새로운 sheet 기본 이름으로 생성
ws.title = "Mysheet"  # sheet  이름 변경
ws.sheet_properties.tabColor = "ffccff"

# Sheet, MySheet, YourSheet
ws1 = wb.create_sheet("YourSheet")  # 주어진 이름으로 Sheet 생성
ws2 = wb.create_sheet("NewSheet", 2)  # 2번째 index 에 sheet 생성

new_ws = wb["NewSheet"]  # Dict  형태로 Sheet에 접근

print(wb.sheetnames)  # 모든 Sheet  이름 확인

#  Sheet 복사
new_ws["A1"] = "Shinjunghun"
target = wb.copy_worksheet(new_ws)
target.title = "CopySheet"

print(wb.sheetnames)

wb.save("sample.xlsx")
