from openpyxl import load_workbook


# 수식을 그대로 가져오는 반복문
wb = load_workbook("sample_formula.xlsx")
ws = wb.active
for row in ws.values:
    for cell in row:
        print(cell)

# 수식이 아닌 데이터를 가져옴
wb = load_workbook("sample_formula.xlsx", data_only=True)
ws = wb.active
for row in ws.values:
    for cell in row:
        print(cell)
